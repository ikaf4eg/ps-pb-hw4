from collections import defaultdict, Counter
import pandas
from openpyxl import load_workbook

# Создаём функцию для подсчёта 'operator' в списке 'c_list'
# border - число, граничное условие.
# compare_list - список для сравнения.
def count_func(operator, c_list, border=0, compare_list = None):
    counter_list = []
    counter_dict = defaultdict(int)
    if operator == 'sales':
        jj = 0
        while jj < border:
            for element in c_list:
                if element == compare_list[jj]:
                    counter_dict[element] += 1
            jj += 1
        counter_list = counter_dict
    else:

        if border == 0 and compare_list == None:
            for element in c_list:
                counter_dict[element[f'{operator}']] += 1
            counter_list = Counter(counter_dict)
        elif border != 0 and compare_list == None:
            jj = 0
            while jj < border:
                for element in c_list[jj][f'{operator}']:
                    counter_dict[element] += 1
                jj += 1
            counter_list = Counter(counter_dict)
        else:
            jj = 0
            while jj < border:
                for element in c_list:
                    if element[f'{operator}'] == compare_list[jj]:
                        counter_dict[element[f'{operator}']] += 1
                jj += 1
            counter_list = counter_dict
    return counter_list


# читаем Логи и создаём словарь
logs_data = pandas.read_excel('logs.xlsx', sheet_name='log', engine='openpyxl')
logs_data_dict = logs_data.to_dict('records')

# Создаём список товаров, разделяем его по запятым и удаляем лишние элементы
sales_list = []
for elements in logs_data_dict:
    item_list = []
    # Определяем структуру записи словаря по продажам
    sales_dict = {
    'gender': '',
    'month': 0,
    'items': [],
    }
    sales_dict['gender'] = elements['Пол']
    sales_dict['month'] = elements['Дата посещения'].month
    item_list.extend(elements['Купленные товары'].split(','))
    for del_sales in item_list:
        if del_sales == 'Ещё 2 варианта' or del_sales == 'Ещё 3 варианта':
            item_list.remove(del_sales)
    sales_dict['items'] = item_list
    sales_list.append(sales_dict)

# Создаём массивы продаж по гендерному признаку
gend_m_list = []
gend_f_list = []
for element in sales_list:
    if element['gender'] == 'м':
        gend_m_list.append(element)
    else:
        gend_f_list.append(element)

# Считаем количество товаров по гендерному признаку по количеству продаж
sales_m_counter = count_func('items', gend_m_list, len(gend_m_list))
sales_f_counter = count_func('items', gend_f_list, len(gend_f_list))

# определяем количество месяцев в логе
max_month = 1
for element in logs_data_dict:
    if element['Дата посещения'].month > max_month:
        max_month = element['Дата посещения'].month

# Ищем самые популярные товары
sales_dict = defaultdict(int)
count_sales = 0
while count_sales < len(sales_list):
    for element in sales_list[count_sales]['items']:
        sales_dict[element] += 1
    count_sales += 1
sales_counter = Counter(sales_dict)

# Создаём список названий самых популярных товаров
trend_sales = []
i = 0
while i < 7:
    trend_sales.append(sales_counter.most_common(7)[i][0])
    i += 1

# Разбиваем статистику товаров по месяцам
sales_list_month = []
month = 1
while month <= max_month:
    sales_month_temp = []
    for element in sales_list:
        if element['month'] == month:
            sales_month_temp += element['items']
    sales_list_month.append(sales_month_temp)
    month +=1

# Создаём словарь по месяцам
sales_list_month_calc = []
month = 1
while month <= max_month:
    sales_list_month_calc.append(count_func('sales',sales_list_month[month-1],7,trend_sales))
    month += 1

# Ищем самый популярный браузер
browser_counter = count_func('Браузер', logs_data_dict)

# Создаём список названий самых популярных браузеров
trend_browser = []
i = 0
while i < 7:
    trend_browser.append(browser_counter.most_common(7)[i][0])
    i += 1

# Разбиваем статистику браузеров по месяцам
browser_list_month = []
month = 1
while month <= max_month:
    browser_month_temp = []
    for element in logs_data_dict:
        if element['Дата посещения'].month == month:
            browser_month_temp.append(element)
    browser_list_month.append(browser_month_temp)
    month +=1

# Создаём словарь по месяцам
browser_list_month_calc = []
month = 1
while month <= max_month:
    browser_list_month_calc.append(count_func('Браузер', browser_list_month[month-1],7,trend_browser))
    month += 1

# выводим полученные данные в файл report.xlsx
wb = load_workbook(filename='report.xlsx')
sheet = wb['Лист1'];
i = 0
while i < 7:
    m = 1
    sheet.cell(row=i+5, column=1).value = browser_counter.most_common(7)[i][0]
    sheet.cell(row=i+19, column=1).value = sales_counter.most_common(7)[i][0]
    while m <= max_month:
        sheet.cell(row=i+5, column=m+2).value = browser_list_month_calc[m-1][trend_browser[i]]
        sheet.cell(row=i+19, column=m+2).value = sales_list_month_calc[m-1][trend_sales[i]]
        m += 1
    i += 1
sheet.cell(row=31, column=2).value = f"{sales_m_counter.most_common(1)[0][0]}"
sheet.cell(row=32, column=2).value = f"{sales_f_counter.most_common(1)[0][0]}"
sheet.cell(row=33, column=2).value = f"{sales_m_counter.most_common()[len(sales_m_counter)-1][0]}"
sheet.cell(row=34, column=2).value = f"{sales_f_counter.most_common()[len(sales_f_counter)-1][0]}"
wb.save('report.xlsx')